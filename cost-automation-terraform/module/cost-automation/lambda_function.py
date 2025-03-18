import json
import logging
import os
import time
import io
import boto3
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from botocore.exceptions import NoCredentialsError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import traceback
import re

from io import BytesIO
permition= False

# Configuration from environment variables
ATHENA_QUERY_Detailed = os.environ.get('ATHENA_QUERY_Detailed')
ATHENA_QUERY_Summary = os.environ.get('ATHENA_QUERY_Summary')
BUCKET_NAME = os.environ.get('BUCKET_NAME')
CC_RECIPIENT = os.environ.get('CC_RECIPIENT')
OUTPUT_LOCATION = os.environ.get('OUTPUT_LOCATION')
PRIMARY_RECIPIENT = os.environ.get('PRIMARY_RECIPIENT')
REGION = os.environ.get('REGION')
SENDER = os.environ.get('SENDER')

recipient_name_match = re.match(r'([a-zA-Z0-9._%+-]+)@([a-zA-Z0-9.-]+)\.([a-zA-Z]{2,})', PRIMARY_RECIPIENT)
if recipient_name_match:
    recipient_name = recipient_name_match.group(1).replace('.', ' ')
else:
    recipient_name = "Recipient"
    
MESSAGE = f"Dear {recipient_name},\nAttached herewith is the cost allocation report for all AWS accounts covering the previous month in Excel format.\nShould you have any questions or require further information, please do not hesitate to contact us. We are at your disposal to assist.\n\nBest regards,\nYour AWS Foundation Team"
#MESSAGE= f"Dear {PRIMARY_RECIPIENT},\nAttached herewith is the cost allocation report for all AWS accounts covering the previous month in Excel format.\nShould you have any questions or require further information, please do not hesitate to contact us. We are at your disposal to assist.\n\nBest regards,\nYour AWS Foundation Team"

# Error Notification with SES
ERROR_SUBJECT = os.environ.get('ERROR_SUBJECT')
PRIMARY_ERROR_RECIPIENT = os.environ.get('PRIMARY_ERROR_RECIPIENT')
CC_ERROR_RECIPIENT = os.environ.get('CC_ERROR_RECIPIENT')


def send_email_error_notification(error_message):
    # Create SES client
    ses_client = boto3.client('ses', region_name=REGION)
    # Send email
    msg = MIMEMultipart()
    msg['Subject'] = ERROR_SUBJECT
    msg['From'] = SENDER
    msg['To'] = PRIMARY_ERROR_RECIPIENT
    msg['Reply-To'] = SENDER
    msg.attach(MIMEText(error_message))

    try:
        response = ses_client.send_raw_email(
            Source=SENDER,
            Destinations=[PRIMARY_ERROR_RECIPIENT],
            RawMessage={'Data': msg.as_string()}
        )
        print(response)
    except Exception as e:
        print("Error sending the error message via email:", e)
        
# Lambda handler function
def execute_athena_query(query_string, filename):
    try:
        # Start the Athena query execution
        athena_client = boto3.client("athena")
        response = athena_client.start_query_execution(
            QueryString=query_string,
            ResultConfiguration={'OutputLocation': OUTPUT_LOCATION}
        )
        print("Athena Query Execution Response:", response)

        # Wait until the query is completed
        query_execution_id = response['QueryExecutionId']
        status = 'RUNNING'
        while status in ['RUNNING', 'QUEUED']:
            try:
                query_execution_info = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
                status = query_execution_info['QueryExecution']['Status']['State']
                print("Query Execution ID:", query_execution_id)
                print("Query Status:", status)

                if 'StateChangeReason' in query_execution_info['QueryExecution']['Status']:
                    print("State Change Reason:", query_execution_info['QueryExecution']['Status']['StateChangeReason'])
            except Exception as e:
                error_message = "Error querying the query status:\n" + str(e)
                print(error_message)
                send_email_error_notification(error_message)
                return

            time.sleep(5)  # Wait for 5 seconds before checking the status again
            
        if status == 'SUCCEEDED':
            # Query was completed successfully
            result_response = athena_client.get_query_results(QueryExecutionId=query_execution_id)
            print('Result response of Athena Query:' + json.dumps(result_response))
            permition = True
        else:
            error_message = "The query failed or has a different status: " + status
            print(error_message)
            send_email_error_notification(error_message)
            return    
        # Get the filename from the output path
        query_execution_info = athena_client.get_query_execution(QueryExecutionId=query_execution_id)
        output_location = query_execution_info['QueryExecution']['ResultConfiguration']['OutputLocation']
        result_file_name = output_location.split("/")[-1] 
        print(query_execution_info)
        print(result_file_name)

        now = datetime.now() - relativedelta(months=1)
        month = str(now.month).zfill(2)
        day = str(now.day).zfill(2)
        date_string = f"{now.year}-{month}-{day}"
        file_name = f"{filename}_{date_string}.csv"
        betrag_summe_rounded = None
        summary_file_name = f"Summary_{date_string}.csv"
        detailed_file_name = f"Detailed_{date_string}.csv"
        
        # Rename the file
        try:
            s3_client = boto3.client("s3")
            s3_client.copy_object(
                Bucket=BUCKET_NAME,
                CopySource={
                    'Bucket': BUCKET_NAME,
                    'Key': f'sahel/{result_file_name}'
                },
                Key=f'sahel/{file_name}',
                MetadataDirective='COPY',
                ContentType='text/csv; charset=utf-8'
            )
            
        except Exception as e:
            error_message = "Error when renaming the generated cost automation file in S3 bucket:\n" + str(e)
            print(error_message)
            send_email_error_notification(error_message)
            return
        try:
            s3_client.upload_file(file_name, BUCKET_NAME, f'sahel/{file_name}')
        except Exception as e:
            error_message=f"Error uploading:{file_name} to the bucket{BUCKET_NAME}/sahel"
        
        try:
            s3_client.delete_object(Bucket=BUCKET_NAME, Key=f'sahel/{result_file_name}.metadata')
            print(f"The metadata file: {result_file_name}.metadata is deleted from S3.")
            s3_client.delete_object(Bucket=BUCKET_NAME, Key=f'sahel/{result_file_name}')
            print(f"The metadata file: {result_file_name}is deleted from S3.")
        except Exception as e:
            error_message = "Error deleting the metadata file in S3 bucket:\n" + str(e)
            print(error_message)
            send_email_error_notification(error_message)    
        try:    
            workbook = Workbook()
            # Erstes Arbeitsblatt für die erste CSV-Datei
            summary_sheet = workbook.create_sheet(title='Summary')
        
             # Daten aus der ersten CSV-Datei hinzufügen (für Summary sheet)
            if permition is True:
                csv_s3_object_summary = s3_client.get_object(Bucket=BUCKET_NAME, Key=f'sahel/{summary_file_name}')
                csv_data_summary = csv_s3_object_summary['Body'].read().decode('utf-8')
        
                csv_reader_summary = csv.reader(io.StringIO(csv_data_summary))
                for row in csv_reader_summary:
                    summary_sheet.append(row)
        
            # Erstes Arbeitsblatt für die erste CSV-Datei
            detailed_sheet = workbook.create_sheet(title='Detailed')
            # Daten aus der zweiten CSV-Datei hinzufügen (für Detailed sheet)
            if permition is True:
                csv_s3_object_detailed = s3_client.get_object(Bucket=BUCKET_NAME, Key=f'sahel/{detailed_file_name}')
                csv_data_detailed = csv_s3_object_detailed['Body'].read().decode('utf-8')
        
                csv_reader_detailed = csv.reader(io.StringIO(csv_data_detailed))
                for row in csv_reader_detailed:
                    detailed_sheet.append(row)
        
            # Speichern und Hochladen der Excel-Datei
            # Aktuelles Datum
            now = datetime.now()
        
            # Datum des vorherigen Monats berechnen
            previous_month = now - timedelta(days=now.day)
            previous_month = previous_month.replace(day=1)  # Erster Tag des vorherigen Monats
        
            # Monatsnamen
            month_names = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        
            # Dateiname für den vorherigen Monat generieren
            excel_file_name = f"{previous_month.year}{previous_month.month:02d}-{month_names[previous_month.month - 1]}-AWSCost.xlsx"
            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)
            s3_client.upload_fileobj(excel_file, BUCKET_NAME, f'sahel/{excel_file_name}')
            print(f"The combined Excel file {excel_file_name} is uploaded to S3.")
            
            # Löschen der CSV- und Metadatendateien
            try:
                s3_client.delete_object(Bucket=BUCKET_NAME, Key=f'sahel/{result_file_name}.metadata')
                print(f"The metadata file: {result_file_name}.metadata is deleted from S3.")
                s3_client.delete_object(Bucket=BUCKET_NAME, Key=f'sahel/{result_file_name}')
                print(f"The metadata file: {result_file_name}is deleted from S3.")
            except Exception as e:
                error_message = "Error deleting the metadata file in S3 bucket:\n" + str(e)
                print(error_message)
                send_email_error_notification(error_message)   
            return excel_file_name
            
        except Exception as e:
            error_message = "Error Accessing result csv files:\n" + str(e)
            print(error_message)
            traceback.print_exc()
            send_email_error_notification(error_message)
            return


    except Exception as e:
        error_message = "Error executing Athena query:\n" + str(e)
        print(error_message)
        send_email_error_notification(error_message)
        return

def execute_athena_queries(query_strings, filenames):
    for query_string, filename in zip(query_strings, filenames):
        excel_file_name = execute_athena_query(query_string, filename)
        if excel_file_name:
            print("Excel File Name:", excel_file_name)
            return excel_file_name  # Gib den Dateinamen als einzelnen String zurück
    print("Excel File Names were found.")
    return None

def lambda_handler(event, context):
    filename_Detailed = "Detailed"
    filename_Summary = "Summary"
    
    try:
        excel_file_name = execute_athena_queries([ATHENA_QUERY_Summary, ATHENA_QUERY_Detailed], [filename_Summary, filename_Detailed])
        
        s3_client = boto3.client('s3')
        try:
            print(f"Trying to retrieve the finalized reporting excel file: {excel_file_name} from S3")
            s3_object = s3_client.get_object(Bucket=BUCKET_NAME, Key=f'sahel/{excel_file_name}')
        except Exception as e:
            print(f"Error retrieving the finalized reporting from S3: {e}")
            return

        # Create SES client
        ses_client = boto3.client('ses', region_name=REGION)

        # Send email with attachment
        msg = MIMEMultipart()
        msg['Subject'] = excel_file_name
        msg['From'] = SENDER
        msg['To'] = PRIMARY_RECIPIENT
        msg['Reply-To'] = SENDER
        msg.attach(MIMEText(MESSAGE))

        if s3_object is not None:
            # Add file content as attachment
            attach = MIMEApplication(s3_object['Body'].read(), _subtype="xlsx")
            attach.add_header('Content-Disposition', 'attachment', filename=excel_file_name)
            msg.attach(attach)
        else:
            print("S3 Object not retrieved successfully. Attachment not added.")

        try:
            response = ses_client.send_raw_email(
                Source=SENDER,
                Destinations=[PRIMARY_RECIPIENT],
                RawMessage={'Data': msg.as_string()}
            )
            print("SES Email sent successfully. Response:", response)
        except NoCredentialsError as e:
            error_message = "Error: Email could not send (NoCredentialsError):\n" + str(e)
            print(error_message)
            send_email_error_notification(error_message)
            return
        except Exception as e:
            error_message = "Error sending SES email:\n" + str(e)
            print(error_message)
            send_email_error_notification(error_message)
            return
    except Exception as e:
        print(f"Error executing Athena queries: {e}")
        # Handle the error as needed
        return
